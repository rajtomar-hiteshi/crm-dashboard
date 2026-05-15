"""
Incremental sync service — downloads Google Sheets, compares with DB,
inserts only NEW rows. Never deletes or updates existing data.
"""
import os, re, json, logging, io
from datetime import datetime, date
from pathlib import Path

import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# ── Config ──────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
CREDS_FILE = PROJECT_ROOT / "credentials.json"
DOWNLOAD_DIR = PROJECT_ROOT / "drive_downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)


# No hardcoded persons/files — sync queries the database dynamically.

# ── Dedup match columns per table ───────────────────
MATCH_COLS = {
    "target_tracking":     ["person_id", "activity_date", "source_file_id"],
    "linkedin_connections": ["person_id", "activity_date", "client_linkedin_url", "source_file_id"],
    "linkedin_followups":  ["person_id", "activity_date", "client_linkedin_url", "follow_up_type", "source_file_id"],
    "linkedin_inmails":    ["person_id", "activity_date", "client_linkedin_url", "source_file_id"],
    "emails":              ["person_id", "activity_date", "client_name", "source_file_id"],
    "positive_responses":  ["person_id", "client_name", "client_linkedin_id", "source_file_id"],
    "leads_generated":     ["person_id", "client_name", "company_name", "source_file_id"],
    "data_extraction":     ["person_id", "activity_date", "client_linkedin_url", "source_file_id"],
    "biddetail_tenders":   ["person_id", "tender_no", "source_file_id"],
    "other_worksheet_data": ["person_id", "original_worksheet", "row_number", "source_file_id"],
}

# ── Column mappings (from ingest_drive.py) ──────────
def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip().lower())

TARGET_TRACKING_MAP = {
    'date': 'activity_date', 'ishaan.mirchandani@hiteshi.com': 'activity_date',
    'linkedin connections': 'linkedin_connections', 'linkedin follow ups': 'linkedin_follow_ups',
    'linkedin follow up': 'linkedin_follow_ups', 'linkedin inmails': 'linkedin_inmails',
    'emails': 'emails', 'data extraction': 'data_extraction', 'cold calling': 'cold_calling',
    'follow up calls': 'follow_up_calls', 'comment': 'comments', 'comments': 'comments',
    'positive responses': 'positive_responses', 'positive response': 'positive_responses',
    'lead generated': 'leads_generated', 'leads generated': 'leads_generated',
    'calls': 'calls', 'team member name': 'team_member_name',
    'planned linkedin connections': 'planned_linkedin_connections',
    'achieved linkedin connections': 'achieved_linkedin_connections',
    'planned linkedin follow ups': 'planned_linkedin_follow_ups',
    'achieved linkedin follow ups': 'achieved_linkedin_follow_ups',
    'inmails': 'planned_inmails', 'achieved inmails': 'achieved_inmails',
    'achieved emails': 'achieved_emails', 'planned emails': 'planned_emails',
    'achieved data extraction': 'achieved_data_extraction',
    'planned data extraction': 'planned_data_extraction',
    'planned cold calls': 'planned_cold_calls', 'achieved cold calls': 'achieved_cold_calls',
    'planned follow up calls': 'planned_follow_up_calls',
    'achieved follow up calls': 'achieved_follow_up_calls',
    'targeted positive responses': 'targeted_positive_responses',
    'achieved positive responses': 'achieved_positive_responses',
    'pre sales lead generated': 'pre_sales_lead_generated', 'other tasks': 'other_tasks',
}
TARGET_TRACKING_SKIP = {'100.0'}

LINKEDIN_CONNECTIONS_MAP = {
    'date': 'activity_date', 'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url', 'client linkedin url': 'client_linkedin_url',
    'client linkedln url': 'client_linkedin_url', 'linkedin account used': 'linkedin_account_used',
    'linkedin account associated': 'linkedin_account_used', 'requested by': 'linkedin_account_used',
    'connection message': 'connection_message', 'connection request sent': 'connection_message',
    'geography': 'geography', 'company size': 'company_size', 'industry': 'industry',
    'cadence sequence': 'cadence_sequence', 'cadence': 'cadence_sequence',
    'accepted': 'accepted', 'filter link': 'filter_link', 'filter name': 'filter_link',
    'filter': 'filter_link', 'response received': 'response_received',
    'comments': 'comments', 'comment': 'comments',
}

LINKEDIN_FOLLOWUPS_MAP = {
    'date': 'activity_date', 'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url', 'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url', 'linkedin account used': 'linkedin_account_used',
    'requested by': 'linkedin_account_used', 'follow up type': 'follow_up_type',
    'message sent': 'message_sent', 'message': 'message_sent',
    'filter': 'filter', 'cadence': 'cadence', 'response received': 'response_received',
}

LINKEDIN_INMAILS_MAP = {
    'date': 'activity_date', 'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url', 'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url', 'linkedin account used': 'linkedin_account_used',
    'requested by': 'linkedin_account_used', 'inmail message sent': 'inmail_message_sent',
    'inmail sent': 'inmail_message_sent', 'message sent': 'inmail_message_sent',
    'geography': 'geography', 'company size': 'company_size', 'industry': 'industry',
    'filter': 'filter', 'cadence': 'cadence',
}

EMAILS_MAP = {
    'date': 'activity_date', 'lead generation exec': 'lead_generation_exec',
    'client name': 'client_name', 'client full name': 'client_name', "client's name": 'client_name',
    'client email': 'client_email', 'client email address': 'client_email',
    'email address': 'client_email', 'linkedin url': 'client_linkedin_url',
    'linkedin id': 'client_linkedin_url', 'linkeidn': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url', 'company name': 'company_name',
    'email content sent': 'email_content_sent', 'email draft': 'email_content_sent',
    'mail sent': 'email_content_sent', 'opportunity url': 'opportunity_url',
    'contact number': 'contact_number', 'reason': 'reason', 'next step': 'next_step',
    'cadence': 'cadence', 'client whatsapp': 'client_whatsapp',
}

POSITIVE_RESPONSES_MAP = {
    'date': 'response_date', 'revert date': 'response_date', 'date client revert': 'response_date',
    'timestamp': 'response_date', 'client type': 'client_type',
    'connected date': 'connected_date', 'connection date': 'connected_date',
    'connected on linkedln': 'connected_date', 'first connected date': 'connected_date',
    'first connection date': 'connected_date', 'lead established date': 'lead_established_date',
    'lead generated dat': 'lead_established_date', 'client name': 'client_name',
    'client linkedin id': 'client_linkedin_id', 'client linkedln id': 'client_linkedin_id',
    'linkedin id': 'client_linkedin_id', 'linkedin id associated': 'linkedin_id_associated',
    'client email address': 'client_email', 'email address': 'client_email',
    'client contact number': 'client_contact_number', 'first follow up': 'first_follow_up',
    'no. of follow ups taken': 'num_follow_ups_taken', 'no. of follow ups': 'num_follow_ups_taken',
    'no. of gap days': 'num_gap_days',
    'response quality (high quality / generic interest)': 'response_quality',
    'response quality ( high quality/ generic interest': 'response_quality',
    'response quality': 'response_quality', 'client first revert': 'client_first_revert',
    'clinet first revert - whole chat ( if available)': 'client_first_revert',
    'chat': 'chat_summary', 'conversation/chat/call summary': 'chat_summary', 'source': 'source',
}

LEADS_GENERATED_MAP = {
    'inquiry date': 'inquiry_date', 'date': 'inquiry_date', 'date of inquiry': 'inquiry_date',
    'lead gen executive': 'lead_gen_executive', 'lead generation exec': 'lead_gen_executive',
    'client name': 'client_name', 'client location': 'client_location', 'city': 'client_location',
    'company name': 'company_name', 'company': 'company_name', 'company size': 'company_size',
    'client designation': 'client_designation', 'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url', 'client email': 'client_email',
    'client number': 'client_contact_number', 'client contact number': 'client_contact_number',
    'summary': 'summary', 'sales inquiry': 'summary', 'next step': 'next_step',
    'next steps': 'next_step', 'next step (if active)': 'next_step',
    'zoho contact link': 'zoho_contact_link', 'zoho deal link (if any)': 'zoho_deal_link',
    'zoho deal link': 'zoho_deal_link', 'lead source': 'lead_source', 'portal': 'lead_source',
    'account': 'account', 'account name': 'account', 'assigned consultant': 'assigned_consultant',
    'current status active/inactive': 'current_status', 'currently active': 'current_status',
    'current status': 'current_status', 'lost reason': 'lost_reason',
    'meeting done': 'meeting_done', 'prospect qualified': 'prospect_qualified',
    'response': 'response', 'status': 'status',
}

RAGINI_LEADS_POSITIONAL = [
    'inquiry_date', 'lead_gen_executive', 'client_name', 'client_location',
    'company_name', 'company_size', 'client_designation', 'client_linkedin_url',
    'client_email', 'summary', 'next_step', 'inquiry_date',
    'lead_source', 'account', 'current_status',
]

DATA_EXTRACTION_MAP = {
    'date': 'activity_date', 'time': 'extraction_time',
    'lead generation exec': 'lead_generation_exec', 'prospect name': 'prospect_name',
    'client name': 'prospect_name', 'prospect company name': 'prospect_company',
    'company name': 'prospect_company', 'company': 'prospect_company',
    'client email address': 'client_email', 'email id': 'client_email',
    'email id 2': 'client_email_2', 'client linkedin profile url': 'client_linkedin_url',
    'linkedin url': 'client_linkedin_url', 'linkedin id': 'client_linkedin_url',
    'opportunity url': 'opportunity_url', 'source of data': 'source_of_data',
    'region': 'region', 'location': 'region', 'designation': 'designation',
    'industry': 'industry', 'contact no.': 'contact_number', 'number': 'contact_number',
    'connection request': 'connection_request', 'email': 'email_sent',
    'likely usecase': 'likely_usecase',
}

BIDDETAIL_MAP = {
    'serial no': 'serial_no', 'sr no': 'serial_no', 'sr. no': 'serial_no',
    'data fetch date': 'data_fetch_date', 'contract date': 'contract_date',
    'query name': 'query_name', 'link of tender': 'link_of_tender',
    'tender no': 'tender_no', 'amount': 'amount',
    'contact person name': 'contact_person_name', 'company': 'company',
    'contact details': 'contact_details', 'bid details email': 'bid_details_email',
    'linkedin email': 'linkedin_email', 'linkedin': 'linkedin',
}

COLUMN_MAP_REGISTRY = {
    'target_tracking': TARGET_TRACKING_MAP,
    'linkedin_connections': LINKEDIN_CONNECTIONS_MAP,
    'linkedin_followups': LINKEDIN_FOLLOWUPS_MAP,
    'linkedin_inmails': LINKEDIN_INMAILS_MAP,
    'emails': EMAILS_MAP,
    'positive_responses': POSITIVE_RESPONSES_MAP,
    'leads_generated': LEADS_GENERATED_MAP,
    'data_extraction': DATA_EXTRACTION_MAP,
    'biddetail_tenders': BIDDETAIL_MAP,
}

TABLE_COLS = {
    'target_tracking': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'linkedin_connections', 'linkedin_follow_ups', 'linkedin_inmails',
        'emails', 'data_extraction', 'cold_calling', 'follow_up_calls',
        'positive_responses', 'leads_generated', 'calls', 'comments',
        'team_member_name', 'planned_linkedin_connections', 'achieved_linkedin_connections',
        'planned_linkedin_follow_ups', 'achieved_linkedin_follow_ups',
        'planned_inmails', 'achieved_inmails', 'planned_emails', 'achieved_emails',
        'planned_data_extraction', 'achieved_data_extraction',
        'planned_cold_calls', 'achieved_cold_calls',
        'planned_follow_up_calls', 'achieved_follow_up_calls',
        'targeted_positive_responses', 'achieved_positive_responses',
        'pre_sales_lead_generated', 'other_tasks',
    ],
    'linkedin_connections': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'connection_message', 'geography', 'company_size', 'industry',
        'cadence_sequence', 'accepted', 'filter_link', 'response_received', 'comments',
    ],
    'linkedin_followups': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'follow_up_type', 'message_sent', 'filter', 'cadence', 'response_received',
    ],
    'linkedin_inmails': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'inmail_message_sent', 'geography', 'company_size', 'industry', 'filter', 'cadence',
    ],
    'emails': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_name', 'client_email', 'client_linkedin_url',
        'company_name', 'email_content_sent', 'opportunity_url', 'contact_number',
        'reason', 'next_step', 'cadence', 'client_whatsapp',
    ],
    'positive_responses': [
        'person_id', 'source_file_id', 'original_worksheet',
        'response_date', 'response_date_raw',
        'client_type', 'connected_date', 'lead_established_date', 'client_name',
        'client_linkedin_id', 'linkedin_id_associated', 'client_email',
        'client_contact_number', 'first_follow_up', 'num_follow_ups_taken',
        'num_gap_days', 'response_quality', 'client_first_revert', 'chat_summary', 'source',
    ],
    'leads_generated': [
        'person_id', 'source_file_id', 'original_worksheet',
        'inquiry_date', 'inquiry_date_raw',
        'lead_gen_executive', 'client_name', 'client_location', 'company_name',
        'company_size', 'client_designation', 'client_linkedin_url', 'client_email',
        'client_contact_number', 'summary', 'next_step', 'zoho_contact_link',
        'zoho_deal_link', 'lead_source', 'account', 'assigned_consultant',
        'current_status', 'lost_reason', 'meeting_done', 'prospect_qualified',
        'response', 'status',
    ],
    'data_extraction': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw', 'extraction_time',
        'lead_generation_exec', 'prospect_name', 'prospect_company',
        'client_email', 'client_email_2', 'client_linkedin_url',
        'opportunity_url', 'source_of_data', 'region', 'designation',
        'industry', 'contact_number', 'connection_request', 'email_sent', 'likely_usecase',
    ],
    'biddetail_tenders': [
        'person_id', 'source_file_id', 'original_worksheet',
        'serial_no', 'data_fetch_date', 'contract_date', 'query_name',
        'link_of_tender', 'tender_no', 'amount', 'contact_person_name',
        'company', 'contact_details', 'bid_details_email', 'linkedin_email', 'linkedin',
    ],
}

DATE_COL = {
    'target_tracking': ('activity_date', 'activity_date_raw'),
    'linkedin_connections': ('activity_date', 'activity_date_raw'),
    'linkedin_followups': ('activity_date', 'activity_date_raw'),
    'linkedin_inmails': ('activity_date', 'activity_date_raw'),
    'emails': ('activity_date', 'activity_date_raw'),
    'positive_responses': ('response_date', 'response_date_raw'),
    'leads_generated': ('inquiry_date', 'inquiry_date_raw'),
    'data_extraction': ('activity_date', 'activity_date_raw'),
}

# ── Helpers ─────────────────────────────────────────
def parse_date(val):
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.date(), str(val)
    if isinstance(val, date):
        return val, str(val)
    raw = str(val).strip()
    if not raw:
        return None, None
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%y',
                '%B %d, %Y', '%B %d,%Y', '%b %d, %Y', '%b %d,%Y',
                '%d %B %Y', '%d %b %Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(raw, fmt).date(), raw
        except ValueError:
            continue
    return None, raw


def cell_to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ('', 'leave', 'holiday', 'n/a', 'na', 'none', '-'):
        return None
    return s


def route_worksheet(ws_name):
    n = ws_name.lower().strip()
    if ('target' in n or 'traget' in n) and 'tracker' not in n and n != 'targets':
        return 'target_tracking'
    if 'connection' in n and ('linkedin' in n or 'linkedln' in n):
        return 'linkedin_connections'
    if 'follow' in n and ('linkedin' in n or 'linkedln' in n) and 'follow up calls' not in n:
        return 'linkedin_followups'
    if 'inmail' in n:
        return 'linkedin_inmails'
    if 'email' in n and 'bid details email' not in n and 'bid detail' not in n:
        return 'emails'
    if 'positive' in n or n in ('pr', 'old pr', 'new pr'):
        return 'positive_responses'
    if 'lead' in n or 'sales inquiry' in n or 'presales' in n or 'deals' in n or 'hot lead' in n:
        return 'leads_generated'
    if 'data extraction' in n or 'data mining' in n:
        return 'data_extraction'
    if 'biddetail' in n or 'tender' in n:
        return 'biddetail_tenders'
    return 'other_worksheet_data'


def is_single_column_tracker(headers):
    if len(headers) <= 1:
        return True
    non_empty = [h for h in headers if h and str(h).strip()]
    return len(non_empty) <= 1


def is_ragini_leads_no_header(ws_name, person_name, headers):
    if 'Ragini' not in person_name:
        return False
    if 'lead' not in ws_name.lower():
        return False
    if headers and len(headers) > 0:
        first = str(headers[0]).strip() if headers[0] else ''
        if re.match(r'\d{4}-\d{2}-\d{2}', first) or re.match(r'\d{2}-\d{2}-\d{4}', first):
            return True
    return False


def read_worksheet_data(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0])
    data_rows = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        data_rows.append(list(row))
    return headers, data_rows


def map_row(headers, row, col_map, skip_cols=None):
    result = {}
    skip = skip_cols or set()
    for i, header in enumerate(headers):
        if i >= len(row):
            break
        h = norm(header)
        if h in skip:
            continue
        db_col = col_map.get(h)
        if db_col:
            val = cell_to_str(row[i])
            if db_col not in result or (result[db_col] is None and val is not None):
                result[db_col] = val
    return result


def map_row_positional(row, positional_map):
    result = {}
    for i, db_col in enumerate(positional_map):
        if i >= len(row):
            break
        if db_col in result and result[db_col] is not None:
            continue
        result[db_col] = cell_to_str(row[i])
    return result


# ── Google Drive ────────────────────────────────────
def get_drive_service():
    creds = Credentials.from_service_account_file(str(CREDS_FILE), scopes=[
        'https://www.googleapis.com/auth/drive.readonly'
    ])
    return build('drive', 'v3', credentials=creds)


def download_file(service, file_id, dest_path):
    file_meta = service.files().get(fileId=file_id, fields='mimeType,name').execute()
    mime_type = file_meta.get('mimeType', '')

    if mime_type == 'application/vnd.google-apps.spreadsheet':
        request = service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        request = service.files().get_media(fileId=file_id)

    with open(dest_path, 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return dest_path


# ── Fingerprint loading ────────────────────────────
def _fingerprint_val(v):
    """Normalize a value for fingerprint comparison — coerce to comparable string."""
    if v is None:
        return ""
    s = str(v).strip()
    if isinstance(v, date):
        return str(v)
    return s


def load_existing_fingerprints(db: Session, table: str, source_file_id: int) -> set:
    """Load existing (match-column) tuples from DB for a given table+source_file."""
    match_cols = MATCH_COLS.get(table)
    if not match_cols:
        return set()
    col_expr = ", ".join(match_cols)
    rows = db.execute(
        text(f"SELECT {col_expr} FROM {table} WHERE source_file_id = :sfid"),
        {"sfid": source_file_id}
    ).fetchall()
    return {tuple(_fingerprint_val(v) for v in row) for row in rows}


def make_fingerprint(mapped: dict, table: str, person_id: int, source_file_id: int) -> tuple:
    """Build a fingerprint tuple from a mapped row dict."""
    match_cols = MATCH_COLS[table]
    vals = []
    for col in match_cols:
        if col == "person_id":
            vals.append(str(person_id))
        elif col == "source_file_id":
            vals.append(str(source_file_id))
        else:
            v = mapped.get(col)
            vals.append(_fingerprint_val(v))
    return tuple(vals)


# ── Core sync logic (shared by incremental + reingest) ──
def _process_file(db, service, sf_row, person_row, incremental=True):
    """Download and process a single source file. Returns (new_rows, skipped_rows, status)."""
    person_id = person_row.id
    sf_id = sf_row.id
    person_name = person_row.full_name
    short_name = person_row.short_name or person_name.split()[0]
    drive_id = sf_row.drive_file_id

    safe_name = re.sub(r'[<>:"/\\|?*]', '_', sf_row.file_name or drive_id)
    if not safe_name.endswith('.xlsx'):
        safe_name += '.xlsx'
    dest = DOWNLOAD_DIR / safe_name

    try:
        download_file(service, drive_id, str(dest))
    except Exception as e:
        logger.error(f"Failed to download {sf_row.file_name}: {e}")
        return 0, 0, "error", {}

    try:
        wb = openpyxl.load_workbook(str(dest), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Error opening {dest}: {e}")
        return 0, 0, "error", {}

    db.execute(
        text("UPDATE source_files SET total_worksheets = :tw, ingested_at = NOW() WHERE id = :sfid"),
        {"tw": len(wb.sheetnames), "sfid": sf_id}
    )

    total_new = 0
    total_skip = 0
    tables_touched = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers_raw, data_rows = read_worksheet_data(ws)
        target_table = route_worksheet(ws_name)

        if not data_rows:
            continue

        if is_single_column_tracker(headers_raw):
            target_table = 'other_worksheet_data'

        if target_table == 'other_worksheet_data':
            existing_fps = load_existing_fingerprints(db, target_table, sf_id) if incremental else set()
            ws_new = 0
            for row_num, row in enumerate(data_rows, start=2):
                if incremental:
                    fp = (str(person_id), ws_name, str(row_num), str(sf_id))
                    if fp in existing_fps:
                        total_skip += 1
                        continue
                row_dict = {}
                for i, val in enumerate(row):
                    key = str(headers_raw[i]).strip() if i < len(headers_raw) and headers_raw[i] else f'col_{i}'
                    if val is not None and str(val).strip():
                        row_dict[key] = str(val).strip()
                if row_dict:
                    db.execute(
                        text("INSERT INTO other_worksheet_data (person_id, source_file_id, original_worksheet, row_number, row_data) VALUES (:pid, :sfid, :ws, :rn, :rd)"),
                        {"pid": person_id, "sfid": sf_id, "ws": ws_name, "rn": row_num, "rd": json.dumps(row_dict)}
                    )
                    ws_new += 1
            total_new += ws_new
            tables_touched['other_worksheet_data'] = tables_touched.get('other_worksheet_data', 0) + ws_new
            if ws_new > 0:
                db.commit()
            continue

        col_map = COLUMN_MAP_REGISTRY.get(target_table, {})
        use_positional = is_ragini_leads_no_header(ws_name, person_name, headers_raw)
        if use_positional:
            data_rows = [list(headers_raw)] + data_rows

        existing_fps = load_existing_fingerprints(db, target_table, sf_id) if incremental else set()
        cols = TABLE_COLS[target_table]
        date_col_info = DATE_COL.get(target_table)
        ws_new = 0
        batch = []

        for row in data_rows:
            if use_positional:
                mapped = map_row_positional(row, RAGINI_LEADS_POSITIONAL)
            else:
                skip = TARGET_TRACKING_SKIP if target_table == 'target_tracking' else set()
                mapped = map_row(headers_raw, row, col_map, skip_cols=skip)

            if not mapped or all(v is None for v in mapped.values()):
                continue

            if date_col_info:
                date_col, raw_col = date_col_info
                raw_val = mapped.get(date_col)
                if raw_val is not None:
                    parsed_dt, raw_str = parse_date(raw_val)
                    mapped[date_col] = str(parsed_dt) if parsed_dt else None
                    mapped[raw_col] = raw_str
                else:
                    mapped[raw_col] = None
                if mapped.get(date_col) is None:
                    continue

            mapped['person_id'] = person_id
            mapped['source_file_id'] = sf_id
            mapped['original_worksheet'] = ws_name

            if incremental:
                fp = make_fingerprint(mapped, target_table, person_id, sf_id)
                if fp in existing_fps:
                    total_skip += 1
                    continue
                existing_fps.add(fp)

            values = tuple(mapped.get(c) for c in cols)
            batch.append(values)
            ws_new += 1

            if len(batch) >= 500:
                _batch_insert(db, target_table, cols, batch)
                batch = []

        if batch:
            _batch_insert(db, target_table, cols, batch)

        if ws_new > 0:
            logger.info(f"  → {target_table}: +{ws_new} new rows from '{ws_name}'")
            tables_touched[target_table] = tables_touched.get(target_table, 0) + ws_new
            db.commit()

        total_new += ws_new

        db.execute(
            text("""INSERT INTO ingestion_log
                    (source_file_id, person_id, worksheet_name, target_table,
                     rows_in_sheet, rows_inserted, columns_in_sheet, columns_mapped, status, ingested_at)
                    VALUES (:sfid, :pid, :ws, :tt, :ris, :ri, :cis, :cm, :st, NOW())"""),
            {"sfid": sf_id, "pid": person_id, "ws": ws_name, "tt": target_table,
             "ris": len(data_rows), "ri": ws_new, "cis": len(headers_raw),
             "cm": len([h for h in headers_raw if norm(h) in col_map]), "st": "SUCCESS"}
        )
        db.commit()

    wb.close()
    return total_new, total_skip, "success", tables_touched


# ── Main sync orchestrator (dynamic — queries DB) ──
def run_incremental_sync(db: Session) -> dict:
    started = datetime.utcnow()
    logger.info("Incremental sync started")

    current_files = db.execute(
        text("""SELECT sf.id, sf.person_id, sf.file_name, sf.drive_file_id, sf.file_type,
                       p.full_name, p.short_name
                FROM source_files sf
                JOIN persons p ON p.id = sf.person_id
                WHERE sf.file_type = 'CURRENT'
                ORDER BY p.full_name""")
    ).fetchall()

    if not current_files:
        return {"status": "success", "message": "No CURRENT files found to sync",
                "new_rows_added": 0, "files_synced": 0, "details": []}

    logger.info(f"Found {len(current_files)} CURRENT files to sync")
    service = get_drive_service()

    total_new = 0
    total_skipped = 0
    files_synced = 0
    person_details = {}
    tables_synced = {}

    for row in current_files:
        class SF:
            id = row[0]; person_id = row[1]; file_name = row[2]; drive_file_id = row[3]
        class PR:
            id = row[1]; full_name = row[5]; short_name = row[6]

        short = PR.short_name or PR.full_name.split()[0]
        logger.info(f"Syncing: {PR.full_name} — {SF.file_name}")

        new, skipped, status, tables = _process_file(db, service, SF(), PR(), incremental=True)

        if short not in person_details:
            person_details[short] = {"person": PR.full_name, "new_rows": 0, "skipped": 0, "status": "success"}
        person_details[short]["new_rows"] += new
        person_details[short]["skipped"] += skipped
        if status == "error":
            person_details[short]["status"] = "error"

        for t, c in tables.items():
            tables_synced[t] = tables_synced.get(t, 0) + c

        total_new += new
        total_skipped += skipped
        files_synced += 1

    db.execute(
        text("""INSERT INTO ingestion_log
                (source_file_id, person_id, worksheet_name, target_table,
                 rows_in_sheet, rows_inserted, columns_in_sheet, columns_mapped, status, ingested_at)
                VALUES (NULL, NULL, 'INCREMENTAL_SYNC', 'all', :ris, :ri, 0, 0, 'SUCCESS', NOW())"""),
        {"ris": total_new + total_skipped, "ri": total_new}
    )
    db.commit()

    finished = datetime.utcnow()
    logger.info(f"Sync complete: {total_new} new rows, {total_skipped} skipped")

    return {
        "status": "success",
        "synced_at": finished.isoformat() + "Z",
        "duration_seconds": round((finished - started).total_seconds(), 1),
        "files_synced": files_synced,
        "new_rows_added": total_new,
        "rows_skipped_already_exist": total_skipped,
        "tables_synced": tables_synced,
        "details": sorted(person_details.values(), key=lambda d: d["person"]),
    }


def run_reingest_past_files(db: Session) -> dict:
    """Re-download and re-ingest ALL PAST files. Assumes old data was already deleted."""
    started = datetime.utcnow()
    logger.info("PAST files re-ingestion started")

    past_files = db.execute(
        text("""SELECT sf.id, sf.person_id, sf.file_name, sf.drive_file_id, sf.file_type,
                       p.full_name, p.short_name
                FROM source_files sf
                JOIN persons p ON p.id = sf.person_id
                WHERE sf.file_type = 'PAST'
                ORDER BY p.full_name""")
    ).fetchall()

    if not past_files:
        return {"status": "success", "message": "No PAST files found", "files_processed": 0}

    logger.info(f"Found {len(past_files)} PAST files to re-ingest")
    service = get_drive_service()

    total_new = 0
    files_processed = 0
    person_details = {}
    tables_synced = {}

    for row in past_files:
        class SF:
            id = row[0]; person_id = row[1]; file_name = row[2]; drive_file_id = row[3]
        class PR:
            id = row[1]; full_name = row[5]; short_name = row[6]

        short = PR.short_name or PR.full_name.split()[0]
        logger.info(f"Re-ingesting PAST: {PR.full_name} — {SF.file_name}")

        new, _, status, tables = _process_file(db, service, SF(), PR(), incremental=False)

        if short not in person_details:
            person_details[short] = {"person": PR.full_name, "new_rows": 0, "status": "success"}
        person_details[short]["new_rows"] += new
        if status == "error":
            person_details[short]["status"] = "error"

        for t, c in tables.items():
            tables_synced[t] = tables_synced.get(t, 0) + c

        total_new += new
        files_processed += 1

    finished = datetime.utcnow()
    logger.info(f"PAST re-ingestion complete: {total_new} rows inserted from {files_processed} files")

    return {
        "status": "success",
        "synced_at": finished.isoformat() + "Z",
        "duration_seconds": round((finished - started).total_seconds(), 1),
        "files_processed": files_processed,
        "total_rows_inserted": total_new,
        "tables_synced": tables_synced,
        "details": sorted(person_details.values(), key=lambda d: d["person"]),
    }


def run_reingest_current_files(db: Session) -> dict:
    """Delete current file data and re-ingest from Google Sheets."""
    started = datetime.utcnow()
    logger.info("CURRENT files re-ingestion started")

    current_files = db.execute(
        text("""SELECT sf.id, sf.person_id, sf.file_name, sf.drive_file_id, sf.file_type,
                       p.full_name, p.short_name
                FROM source_files sf
                JOIN persons p ON p.id = sf.person_id
                WHERE sf.file_type = 'CURRENT'
                ORDER BY p.full_name""")
    ).fetchall()

    if not current_files:
        return {"status": "success", "message": "No CURRENT files found", "files_processed": 0}

    sf_ids = [row[0] for row in current_files]
    logger.info(f"Found {len(current_files)} CURRENT files — deleting old data for sf_ids={sf_ids}")

    tables_to_clean = [
        'target_tracking', 'linkedin_connections', 'linkedin_followups',
        'linkedin_inmails', 'emails', 'data_extraction',
        'positive_responses', 'leads_generated', 'other_worksheet_data',
    ]
    for tbl in tables_to_clean:
        db.execute(text(f"DELETE FROM {tbl} WHERE source_file_id = ANY(:ids)"), {"ids": sf_ids})
    db.execute(text("DELETE FROM ingestion_log WHERE source_file_id = ANY(:ids)"), {"ids": sf_ids})
    db.commit()
    logger.info("Old CURRENT data deleted")

    service = get_drive_service()
    total_new = 0
    files_processed = 0
    person_details = {}
    tables_synced = {}

    for row in current_files:
        class SF:
            id = row[0]; person_id = row[1]; file_name = row[2]; drive_file_id = row[3]
        class PR:
            id = row[1]; full_name = row[5]; short_name = row[6]

        short = PR.short_name or PR.full_name.split()[0]
        logger.info(f"Re-ingesting CURRENT: {PR.full_name} — {SF.file_name}")

        new, _, status, tables = _process_file(db, service, SF(), PR(), incremental=False)

        if short not in person_details:
            person_details[short] = {"person": PR.full_name, "new_rows": 0, "status": "success"}
        person_details[short]["new_rows"] += new
        if status == "error":
            person_details[short]["status"] = "error"

        for t, c in tables.items():
            tables_synced[t] = tables_synced.get(t, 0) + c

        total_new += new
        files_processed += 1

    finished = datetime.utcnow()
    logger.info(f"CURRENT re-ingestion complete: {total_new} rows from {files_processed} files")

    return {
        "status": "success",
        "synced_at": finished.isoformat() + "Z",
        "duration_seconds": round((finished - started).total_seconds(), 1),
        "files_processed": files_processed,
        "total_rows_inserted": total_new,
        "tables_synced": tables_synced,
        "details": sorted(person_details.values(), key=lambda d: d["person"]),
    }


def _batch_insert(db: Session, table: str, cols: list, batch: list):
    col_str = ", ".join(cols)
    placeholders = ", ".join([f":c{i}" for i in range(len(cols))])
    for row_vals in batch:
        params = {f"c{i}": v for i, v in enumerate(row_vals)}
        db.execute(text(f"INSERT INTO {table} ({col_str}) VALUES ({placeholders})"), params)
