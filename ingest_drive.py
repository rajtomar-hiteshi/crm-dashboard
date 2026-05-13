"""
Lead Gen CRM — Google Drive Data Extraction & PostgreSQL Ingestion
Downloads xlsx files from Drive, maps all worksheets to normalized schema, inserts into PostgreSQL.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os, io, re, json, random
from datetime import datetime, date
from collections import defaultdict
import psycopg2
from psycopg2.extras import execute_values
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
DB = dict(host='localhost', port=int(os.getenv('DB_PORT', '5432')), dbname='leadgen_crm', user='postgres', password='postgres')
CREDS_FILE = os.path.join(os.path.dirname(__file__), 'credentials.json')
DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), 'drive_downloads')
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

PERSONS = [
    {"full_name": "Karishma Gurnani", "short_name": "Karishma"},
    {"full_name": "Ragini Mahajan",   "short_name": "Ragini"},
    {"full_name": "Yashika Medhe",    "short_name": "Yashika"},
    {"full_name": "Yogita Thakur",    "short_name": "Yogita"},
]

FILES = [
    {"person": "Karishma Gurnani", "drive_id": "13DQIF2A1Tgow0_PVw1IJMQ5XvzA_VeitbVbCN606noU",
     "name": "[ Jan 2 - 21 april 2026 ] Lead Generation- Karishma Gurnani (2026)",
     "file_type": "PAST", "period_start": "2026-01-02", "period_end": "2026-04-21"},

    {"person": "Karishma Gurnani", "drive_id": "13S1qaUWgtuo1fAYpJIUxgG_YKrwnQe2tjlH8L3vAqTo",
     "name": "[ 22 april 2026 - Current ] Lead Generation - Karishma (2026)",
     "file_type": "CURRENT", "period_start": "2026-04-22", "period_end": None},

    {"person": "Ragini Mahajan", "drive_id": "1Tl344h5Sn33cF5VlpZZmzP-a2bk1Bb7IN6-RBGTo_JM",
     "name": "[ 12th jan 2026 - Current ] Lead Generation - Ragini Mahajan 2026",
     "file_type": "CURRENT", "period_start": "2026-01-12", "period_end": None},

    {"person": "Yashika Medhe", "drive_id": "1hf40HUOiqQG5Nk5ANrUlx1bbio4fuzQ-uF0MteZATbc",
     "name": "[ January 2, 2025 - 30 August 2025 ] Lead Generation - LinkedIn - Daily Report Yashika",
     "file_type": "PAST", "period_start": "2025-01-02", "period_end": "2025-08-30"},

    {"person": "Yashika Medhe", "drive_id": "1DJRBExpAdYuQGX-Y_-IRGrEk5xT0ZVMDBQ8-D_YUAYs",
     "name": "[April 4, 2026 - Current] Lead Generation - Yashika Medhe",
     "file_type": "CURRENT", "period_start": "2026-04-04", "period_end": None},

    {"person": "Yogita Thakur", "drive_id": "15wAch41nIISrgOWGxNb8oFEkqu3A1oWxWUXdykm4kb0",
     "name": "[ 6 jan 2026 - 21 april ] Lead Generation - Yogita Thakur (2026)",
     "file_type": "PAST", "period_start": "2026-01-06", "period_end": "2026-04-21"},

    {"person": "Yogita Thakur", "drive_id": "1DZC1kUfZuvuA579_TMRyb7aCVDfe3hK1a5X_YVJbJkE",
     "name": "[ 22 april 2026 - Current ] Lead Generation - Yogita",
     "file_type": "CURRENT", "period_start": "2026-04-22", "period_end": None},
]

# ──────────────────────────────────────────────
# SCHEMA DDL
# ──────────────────────────────────────────────
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS persons (
    id SERIAL PRIMARY KEY,
    full_name TEXT NOT NULL UNIQUE,
    short_name TEXT,
    email TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS source_files (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id),
    file_name TEXT NOT NULL,
    drive_file_id TEXT,
    file_type TEXT CHECK (file_type IN ('PAST', 'CURRENT')),
    period_start DATE,
    period_end DATE,
    total_worksheets INTEGER,
    ingested_at TIMESTAMP DEFAULT NOW(),
    UNIQUE(drive_file_id)
);

CREATE TABLE IF NOT EXISTS target_tracking (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    linkedin_connections TEXT,
    linkedin_follow_ups TEXT,
    linkedin_inmails TEXT,
    emails TEXT,
    data_extraction TEXT,
    cold_calling TEXT,
    follow_up_calls TEXT,
    positive_responses TEXT,
    leads_generated TEXT,
    calls TEXT,
    comments TEXT,
    team_member_name TEXT,
    planned_linkedin_connections TEXT,
    achieved_linkedin_connections TEXT,
    planned_linkedin_follow_ups TEXT,
    achieved_linkedin_follow_ups TEXT,
    planned_inmails TEXT,
    achieved_inmails TEXT,
    planned_emails TEXT,
    achieved_emails TEXT,
    planned_data_extraction TEXT,
    achieved_data_extraction TEXT,
    planned_cold_calls TEXT,
    achieved_cold_calls TEXT,
    planned_follow_up_calls TEXT,
    achieved_follow_up_calls TEXT,
    targeted_positive_responses TEXT,
    achieved_positive_responses TEXT,
    pre_sales_lead_generated TEXT,
    other_tasks TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_tt_person_date ON target_tracking(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_tt_date ON target_tracking(activity_date);
CREATE INDEX IF NOT EXISTS idx_tt_source ON target_tracking(source_file_id);

CREATE TABLE IF NOT EXISTS linkedin_connections (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    lead_generation_exec TEXT,
    client_linkedin_url TEXT,
    linkedin_account_used TEXT,
    connection_message TEXT,
    geography TEXT,
    company_size TEXT,
    industry TEXT,
    cadence_sequence TEXT,
    accepted TEXT,
    filter_link TEXT,
    response_received TEXT,
    comments TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_lc_person_date ON linkedin_connections(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_lc_date ON linkedin_connections(activity_date);
CREATE INDEX IF NOT EXISTS idx_lc_url ON linkedin_connections(client_linkedin_url);
CREATE INDEX IF NOT EXISTS idx_lc_source ON linkedin_connections(source_file_id);

CREATE TABLE IF NOT EXISTS linkedin_followups (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    lead_generation_exec TEXT,
    client_linkedin_url TEXT,
    linkedin_account_used TEXT,
    follow_up_type TEXT,
    message_sent TEXT,
    filter TEXT,
    cadence TEXT,
    response_received TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_lf_person_date ON linkedin_followups(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_lf_date ON linkedin_followups(activity_date);
CREATE INDEX IF NOT EXISTS idx_lf_url ON linkedin_followups(client_linkedin_url);
CREATE INDEX IF NOT EXISTS idx_lf_source ON linkedin_followups(source_file_id);

CREATE TABLE IF NOT EXISTS linkedin_inmails (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    lead_generation_exec TEXT,
    client_linkedin_url TEXT,
    linkedin_account_used TEXT,
    inmail_message_sent TEXT,
    geography TEXT,
    company_size TEXT,
    industry TEXT,
    filter TEXT,
    cadence TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_li_person_date ON linkedin_inmails(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_li_date ON linkedin_inmails(activity_date);
CREATE INDEX IF NOT EXISTS idx_li_url ON linkedin_inmails(client_linkedin_url);
CREATE INDEX IF NOT EXISTS idx_li_source ON linkedin_inmails(source_file_id);

CREATE TABLE IF NOT EXISTS emails (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    lead_generation_exec TEXT,
    client_name TEXT,
    client_email TEXT,
    client_linkedin_url TEXT,
    company_name TEXT,
    email_content_sent TEXT,
    opportunity_url TEXT,
    contact_number TEXT,
    reason TEXT,
    next_step TEXT,
    cadence TEXT,
    client_whatsapp TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_em_person_date ON emails(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_em_date ON emails(activity_date);
CREATE INDEX IF NOT EXISTS idx_em_source ON emails(source_file_id);

CREATE TABLE IF NOT EXISTS positive_responses (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    response_date DATE,
    response_date_raw TEXT,
    client_type TEXT,
    connected_date TEXT,
    lead_established_date TEXT,
    client_name TEXT,
    client_linkedin_id TEXT,
    linkedin_id_associated TEXT,
    client_email TEXT,
    client_contact_number TEXT,
    first_follow_up TEXT,
    num_follow_ups_taken TEXT,
    num_gap_days TEXT,
    response_quality TEXT,
    client_first_revert TEXT,
    chat_summary TEXT,
    source TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_pr_person_date ON positive_responses(person_id, response_date);
CREATE INDEX IF NOT EXISTS idx_pr_date ON positive_responses(response_date);
CREATE INDEX IF NOT EXISTS idx_pr_source ON positive_responses(source_file_id);

CREATE TABLE IF NOT EXISTS leads_generated (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    inquiry_date DATE,
    inquiry_date_raw TEXT,
    lead_gen_executive TEXT,
    client_name TEXT,
    client_location TEXT,
    company_name TEXT,
    company_size TEXT,
    client_designation TEXT,
    client_linkedin_url TEXT,
    client_email TEXT,
    client_contact_number TEXT,
    summary TEXT,
    next_step TEXT,
    zoho_contact_link TEXT,
    zoho_deal_link TEXT,
    lead_source TEXT,
    account TEXT,
    assigned_consultant TEXT,
    current_status TEXT,
    lost_reason TEXT,
    meeting_done TEXT,
    prospect_qualified TEXT,
    response TEXT,
    status TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_lg_person_date ON leads_generated(person_id, inquiry_date);
CREATE INDEX IF NOT EXISTS idx_lg_date ON leads_generated(inquiry_date);
CREATE INDEX IF NOT EXISTS idx_lg_source ON leads_generated(source_file_id);

CREATE TABLE IF NOT EXISTS data_extraction (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    activity_date DATE,
    activity_date_raw TEXT,
    extraction_time TEXT,
    lead_generation_exec TEXT,
    prospect_name TEXT,
    prospect_company TEXT,
    client_email TEXT,
    client_email_2 TEXT,
    client_linkedin_url TEXT,
    opportunity_url TEXT,
    source_of_data TEXT,
    region TEXT,
    designation TEXT,
    industry TEXT,
    contact_number TEXT,
    connection_request TEXT,
    email_sent TEXT,
    likely_usecase TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_de_person_date ON data_extraction(person_id, activity_date);
CREATE INDEX IF NOT EXISTS idx_de_source ON data_extraction(source_file_id);

CREATE TABLE IF NOT EXISTS biddetail_tenders (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    serial_no TEXT,
    data_fetch_date TEXT,
    contract_date TEXT,
    query_name TEXT,
    link_of_tender TEXT,
    tender_no TEXT,
    amount TEXT,
    contact_person_name TEXT,
    company TEXT,
    contact_details TEXT,
    bid_details_email TEXT,
    linkedin_email TEXT,
    linkedin TEXT,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_bt_person ON biddetail_tenders(person_id);
CREATE INDEX IF NOT EXISTS idx_bt_source ON biddetail_tenders(source_file_id);

CREATE TABLE IF NOT EXISTS other_worksheet_data (
    id SERIAL PRIMARY KEY,
    person_id INTEGER REFERENCES persons(id) NOT NULL,
    source_file_id INTEGER REFERENCES source_files(id) NOT NULL,
    original_worksheet TEXT NOT NULL,
    row_number INTEGER,
    row_data JSONB NOT NULL,
    created_at TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_ow_person ON other_worksheet_data(person_id);
CREATE INDEX IF NOT EXISTS idx_ow_worksheet ON other_worksheet_data(original_worksheet);
CREATE INDEX IF NOT EXISTS idx_ow_source ON other_worksheet_data(source_file_id);

CREATE TABLE IF NOT EXISTS ingestion_log (
    id SERIAL PRIMARY KEY,
    source_file_id INTEGER REFERENCES source_files(id),
    person_id INTEGER REFERENCES persons(id),
    worksheet_name TEXT NOT NULL,
    target_table TEXT NOT NULL,
    rows_in_sheet INTEGER NOT NULL,
    rows_inserted INTEGER NOT NULL,
    columns_in_sheet INTEGER NOT NULL,
    columns_mapped INTEGER NOT NULL,
    status TEXT CHECK (status IN ('SUCCESS', 'PARTIAL', 'FAILED', 'SKIPPED_EMPTY')),
    error_message TEXT,
    sample_row_check_passed BOOLEAN,
    numeric_checksum_passed BOOLEAN,
    ingested_at TIMESTAMP DEFAULT NOW()
);
"""

# ──────────────────────────────────────────────
# COLUMN MAPPINGS (source header → db column)
# ──────────────────────────────────────────────
def norm(s):
    """Normalize header: lowercase, collapse whitespace, strip"""
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip().lower())

TARGET_TRACKING_MAP = {
    'date': 'activity_date',
    'ishaan.mirchandani@hiteshi.com': 'activity_date',
    'linkedin connections': 'linkedin_connections',
    'linkedin follow ups': 'linkedin_follow_ups',
    'linkedin follow up': 'linkedin_follow_ups',
    'linkedin inmails': 'linkedin_inmails',
    'emails': 'emails',
    'data extraction': 'data_extraction',
    'cold calling': 'cold_calling',
    'follow up calls': 'follow_up_calls',
    'comment': 'comments',
    'comments': 'comments',
    'positive responses': 'positive_responses',
    'positive response': 'positive_responses',
    'lead generated': 'leads_generated',
    'leads generated': 'leads_generated',
    'calls': 'calls',
    'team member name': 'team_member_name',
    'planned linkedin connections': 'planned_linkedin_connections',
    'achieved linkedin connections': 'achieved_linkedin_connections',
    'planned linkedin follow ups': 'planned_linkedin_follow_ups',
    'achieved linkedin follow ups': 'achieved_linkedin_follow_ups',
    'inmails': 'planned_inmails',
    'achieved inmails': 'achieved_inmails',
    'achieved emails': 'achieved_emails',
    'planned emails': 'planned_emails',
    'achieved data extraction': 'achieved_data_extraction',
    'planned data extraction': 'planned_data_extraction',
    'planned cold calls': 'planned_cold_calls',
    'achieved cold calls': 'achieved_cold_calls',
    'planned follow up calls': 'planned_follow_up_calls',
    'achieved follow up calls': 'achieved_follow_up_calls',
    'targeted positive responses': 'targeted_positive_responses',
    'achieved positive responses': 'achieved_positive_responses',
    'pre sales lead generated': 'pre_sales_lead_generated',
    'other tasks': 'other_tasks',
}
TARGET_TRACKING_SKIP = {'100.0'}

LINKEDIN_CONNECTIONS_MAP = {
    'date': 'activity_date',
    'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url',
    'client linkedln url': 'client_linkedin_url',
    'linkedin account used': 'linkedin_account_used',
    'linkedin account associated': 'linkedin_account_used',
    'requested by': 'linkedin_account_used',
    'connection message': 'connection_message',
    'connection request sent': 'connection_message',
    'geography': 'geography',
    'company size': 'company_size',
    'industry': 'industry',
    'cadence sequence': 'cadence_sequence',
    'cadence': 'cadence_sequence',
    'accepted': 'accepted',
    'filter link': 'filter_link',
    'filter name': 'filter_link',
    'filter': 'filter_link',
    'response received': 'response_received',
    'comments': 'comments',
    'comment': 'comments',
}

LINKEDIN_FOLLOWUPS_MAP = {
    'date': 'activity_date',
    'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url',
    'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url',
    'linkedin account used': 'linkedin_account_used',
    'requested by': 'linkedin_account_used',
    'follow up type': 'follow_up_type',
    'message sent': 'message_sent',
    'message': 'message_sent',
    'filter': 'filter',
    'cadence': 'cadence',
    'response received': 'response_received',
}

LINKEDIN_INMAILS_MAP = {
    'date': 'activity_date',
    'lead generation exec': 'lead_generation_exec',
    'client linkedin profile url': 'client_linkedin_url',
    'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url',
    'linkedin account used': 'linkedin_account_used',
    'requested by': 'linkedin_account_used',
    'inmail message sent': 'inmail_message_sent',
    'inmail sent': 'inmail_message_sent',
    'message sent': 'inmail_message_sent',
    'geography': 'geography',
    'company size': 'company_size',
    'industry': 'industry',
    'filter': 'filter',
    'cadence': 'cadence',
}

EMAILS_MAP = {
    'date': 'activity_date',
    'lead generation exec': 'lead_generation_exec',
    'client name': 'client_name',
    'client full name': 'client_name',
    "client's name": 'client_name',
    'client email': 'client_email',
    'client email address': 'client_email',
    'email address': 'client_email',
    'linkedin url': 'client_linkedin_url',
    'linkedin id': 'client_linkedin_url',
    'linkeidn': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url',
    'company name': 'company_name',
    'email content sent': 'email_content_sent',
    'email draft': 'email_content_sent',
    'mail sent': 'email_content_sent',
    'opportunity url': 'opportunity_url',
    'contact number': 'contact_number',
    'reason': 'reason',
    'next step': 'next_step',
    'cadence': 'cadence',
    'client whatsapp': 'client_whatsapp',
}

POSITIVE_RESPONSES_MAP = {
    'date': 'response_date',
    'revert date': 'response_date',
    'date client revert': 'response_date',
    'timestamp': 'response_date',
    'client type': 'client_type',
    'connected date': 'connected_date',
    'connection date': 'connected_date',
    'connected on linkedln': 'connected_date',
    'first connected date': 'connected_date',
    'first connection date': 'connected_date',
    'lead established date': 'lead_established_date',
    'lead generated dat': 'lead_established_date',
    'client name': 'client_name',
    'client linkedin id': 'client_linkedin_id',
    'client linkedln id': 'client_linkedin_id',
    'linkedin id': 'client_linkedin_id',
    'linkedin id associated': 'linkedin_id_associated',
    'client email address': 'client_email',
    'email address': 'client_email',
    'client contact number': 'client_contact_number',
    'first follow up': 'first_follow_up',
    'no. of follow ups taken': 'num_follow_ups_taken',
    'no. of follow ups': 'num_follow_ups_taken',
    'no. of gap days': 'num_gap_days',
    'response quality (high quality / generic interest)': 'response_quality',
    'response quality ( high quality/ generic interest': 'response_quality',
    'response quality': 'response_quality',
    'client first revert': 'client_first_revert',
    'clinet first revert - whole chat ( if available)': 'client_first_revert',
    'chat': 'chat_summary',
    'conversation/chat/call summary': 'chat_summary',
    'source': 'source',
}

LEADS_GENERATED_MAP = {
    'inquiry date': 'inquiry_date',
    'date': 'inquiry_date',
    'date of inquiry': 'inquiry_date',
    'lead gen executive': 'lead_gen_executive',
    'lead generation exec': 'lead_gen_executive',
    'client name': 'client_name',
    'client location': 'client_location',
    'city': 'client_location',
    'company name': 'company_name',
    'company': 'company_name',
    'company size': 'company_size',
    'client designation': 'client_designation',
    'linkedin url': 'client_linkedin_url',
    'client linkedin url': 'client_linkedin_url',
    'client email': 'client_email',
    'client number': 'client_contact_number',
    'client contact number': 'client_contact_number',
    'summary': 'summary',
    'sales inquiry': 'summary',
    'next step': 'next_step',
    'next steps': 'next_step',
    'next step (if active)': 'next_step',
    'zoho contact link': 'zoho_contact_link',
    'zoho contact link': 'zoho_contact_link',
    'zoho deal link (if any)': 'zoho_deal_link',
    'zoho deal link': 'zoho_deal_link',
    'lead source': 'lead_source',
    'portal': 'lead_source',
    'account': 'account',
    'account name': 'account',
    'assigned consultant': 'assigned_consultant',
    'current status active/inactive': 'current_status',
    'currently active': 'current_status',
    'current status': 'current_status',
    'lost reason': 'lost_reason',
    'meeting done': 'meeting_done',
    'prospect qualified': 'prospect_qualified',
    'response': 'response',
    'status': 'status',
}

RAGINI_LEADS_POSITIONAL = [
    'inquiry_date', 'lead_gen_executive', 'client_name', 'client_location',
    'company_name', 'company_size', 'client_designation', 'client_linkedin_url',
    'client_email', 'summary', 'next_step', 'inquiry_date',  # alt date
    'lead_source', 'account', 'current_status',
]

DATA_EXTRACTION_MAP = {
    'date': 'activity_date',
    'time': 'extraction_time',
    'lead generation exec': 'lead_generation_exec',
    'prospect name': 'prospect_name',
    'client name': 'prospect_name',
    'prospect company name': 'prospect_company',
    'company name': 'prospect_company',
    'company': 'prospect_company',
    'client email address': 'client_email',
    'email id': 'client_email',
    'email id 2': 'client_email_2',
    'client linkedin profile url': 'client_linkedin_url',
    'linkedin url': 'client_linkedin_url',
    'linkedin id': 'client_linkedin_url',
    'opportunity url': 'opportunity_url',
    'source of data': 'source_of_data',
    'region': 'region',
    'location': 'region',
    'designation': 'designation',
    'industry': 'industry',
    'contact no.': 'contact_number',
    'number': 'contact_number',
    'connection request': 'connection_request',
    'email': 'email_sent',
    'likely usecase': 'likely_usecase',
}

BIDDETAIL_MAP = {
    'serial no': 'serial_no',
    'sr no': 'serial_no',
    'sr. no': 'serial_no',
    'data fetch date': 'data_fetch_date',
    'contract date': 'contract_date',
    'query name': 'query_name',
    'link of tender': 'link_of_tender',
    'tender no': 'tender_no',
    'amount': 'amount',
    'contact person name': 'contact_person_name',
    'company': 'company',
    'contact details': 'contact_details',
    'bid details email': 'bid_details_email',
    'linkedin email': 'linkedin_email',
    'linkedin': 'linkedin',
}

# ──────────────────────────────────────────────
# WORKSHEET ROUTING
# ──────────────────────────────────────────────
def route_worksheet(ws_name):
    """Determine target table from worksheet name."""
    n = ws_name.lower().strip()

    if ('target' in n or 'traget' in n) and 'tracker' not in n and n != 'targets':
        return 'target_tracking'
    if 'connection' in n and ('linkedin' in n or 'linkedln' in n):
        return 'linkedin_connections'
    if 'follow' in n and ('linkedin' in n or 'linkedln' in n) and 'follow up calls' not in n:
        return 'linkedin_followups'
    if 'inmail' in n:
        return 'linkedin_inmails'
    if 'email' in n and 'bid details email' not in n and 'bid detail' not in n:
        return 'emails'
    if 'positive' in n or n in ('pr', 'old pr', 'new pr'):
        return 'positive_responses'
    if 'lead' in n or 'sales inquiry' in n or 'presales' in n or 'deals' in n or 'hot lead' in n:
        return 'leads_generated'
    if 'data extraction' in n or 'data mining' in n:
        return 'data_extraction'
    if 'biddetail' in n or 'tender' in n:
        return 'biddetail_tenders'
    return 'other_worksheet_data'

def get_column_map(target_table):
    return {
        'target_tracking': TARGET_TRACKING_MAP,
        'linkedin_connections': LINKEDIN_CONNECTIONS_MAP,
        'linkedin_followups': LINKEDIN_FOLLOWUPS_MAP,
        'linkedin_inmails': LINKEDIN_INMAILS_MAP,
        'emails': EMAILS_MAP,
        'positive_responses': POSITIVE_RESPONSES_MAP,
        'leads_generated': LEADS_GENERATED_MAP,
        'data_extraction': DATA_EXTRACTION_MAP,
        'biddetail_tenders': BIDDETAIL_MAP,
    }.get(target_table, {})

# Date column names per table (for _raw audit)
DATE_COL = {
    'target_tracking': ('activity_date', 'activity_date_raw'),
    'linkedin_connections': ('activity_date', 'activity_date_raw'),
    'linkedin_followups': ('activity_date', 'activity_date_raw'),
    'linkedin_inmails': ('activity_date', 'activity_date_raw'),
    'emails': ('activity_date', 'activity_date_raw'),
    'positive_responses': ('response_date', 'response_date_raw'),
    'leads_generated': ('inquiry_date', 'inquiry_date_raw'),
    'data_extraction': ('activity_date', 'activity_date_raw'),
}

# ──────────────────────────────────────────────
# DATE PARSING
# ──────────────────────────────────────────────
def parse_date(val):
    """Try to parse a date from various formats. Returns (date_obj, raw_str) or (None, raw_str)."""
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.date(), str(val)
    if isinstance(val, date):
        return val, str(val)
    raw = str(val).strip()
    if not raw:
        return None, None

    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%y',
                '%B %d, %Y', '%B %d,%Y', '%b %d, %Y', '%b %d,%Y',
                '%d %B %Y', '%d %b %Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(raw, fmt).date(), raw
        except ValueError:
            continue
    return None, raw

def cell_to_str(val):
    """Convert cell value to string, treating Leave/Holiday/N-A as None."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ('', 'leave', 'holiday', 'n/a', 'na', 'none', '-'):
        return None
    return s

# ──────────────────────────────────────────────
# GOOGLE DRIVE DOWNLOAD
# ──────────────────────────────────────────────
def get_drive_service():
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=[
        'https://www.googleapis.com/auth/drive.readonly'
    ])
    return build('drive', 'v3', credentials=creds)

def download_file(service, file_id, dest_path):
    """Export a Google Sheet as xlsx from Google Drive."""
    request = service.files().export_media(
        fileId=file_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    with open(dest_path, 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return dest_path

# ──────────────────────────────────────────────
# WORKSHEET PROCESSING
# ──────────────────────────────────────────────
def is_single_column_tracker(ws, headers):
    """Check if worksheet is a single-column text dump."""
    if len(headers) <= 1:
        return True
    non_empty = [h for h in headers if h and str(h).strip()]
    if len(non_empty) <= 1:
        return True
    single_text_headers = {
        'linkedin followup records', 'daily linkedin connections tracker',
        'positive response database', 'linkedin connection records',
    }
    if len(non_empty) == 1 and norm(non_empty[0]) in single_text_headers:
        return True
    return False

def is_ragini_leads_no_header(ws_name, person_name, headers):
    """Check if this is Ragini's Lead Generated sheet with data-as-headers."""
    if 'Ragini' not in person_name:
        return False
    n = ws_name.lower()
    if 'lead' not in n:
        return False
    if headers and len(headers) > 0:
        first = str(headers[0]).strip() if headers[0] else ''
        if re.match(r'\d{4}-\d{2}-\d{2}', first) or re.match(r'\d{2}-\d{2}-\d{4}', first):
            return True
    return False

def read_worksheet_data(ws):
    """Read all rows from an openpyxl worksheet. Returns (headers, data_rows)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = list(rows[0])
    data_rows = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        data_rows.append(list(row))
    return headers, data_rows

def map_row(headers, row, col_map, skip_cols=None):
    """Map a single row using column mapping. Returns dict of db_col -> value."""
    result = {}
    skip = skip_cols or set()
    for i, header in enumerate(headers):
        if i >= len(row):
            break
        h = norm(header)
        if h in skip:
            continue
        db_col = col_map.get(h)
        if db_col:
            val = cell_to_str(row[i])
            if db_col not in result or (result[db_col] is None and val is not None):
                result[db_col] = val
    return result

def map_row_positional(row, positional_map):
    """Map row by column position (for Ragini's leads)."""
    result = {}
    for i, db_col in enumerate(positional_map):
        if i >= len(row):
            break
        if db_col in result and result[db_col] is not None:
            continue
        result[db_col] = cell_to_str(row[i])
    return result

# ──────────────────────────────────────────────
# DB COLUMNS PER TABLE (for INSERT)
# ──────────────────────────────────────────────
TABLE_COLS = {
    'target_tracking': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'linkedin_connections', 'linkedin_follow_ups', 'linkedin_inmails',
        'emails', 'data_extraction', 'cold_calling', 'follow_up_calls',
        'positive_responses', 'leads_generated', 'calls', 'comments',
        'team_member_name', 'planned_linkedin_connections', 'achieved_linkedin_connections',
        'planned_linkedin_follow_ups', 'achieved_linkedin_follow_ups',
        'planned_inmails', 'achieved_inmails', 'planned_emails', 'achieved_emails',
        'planned_data_extraction', 'achieved_data_extraction',
        'planned_cold_calls', 'achieved_cold_calls',
        'planned_follow_up_calls', 'achieved_follow_up_calls',
        'targeted_positive_responses', 'achieved_positive_responses',
        'pre_sales_lead_generated', 'other_tasks',
    ],
    'linkedin_connections': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'connection_message', 'geography', 'company_size', 'industry',
        'cadence_sequence', 'accepted', 'filter_link', 'response_received', 'comments',
    ],
    'linkedin_followups': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'follow_up_type', 'message_sent', 'filter', 'cadence', 'response_received',
    ],
    'linkedin_inmails': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_linkedin_url', 'linkedin_account_used',
        'inmail_message_sent', 'geography', 'company_size', 'industry', 'filter', 'cadence',
    ],
    'emails': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw',
        'lead_generation_exec', 'client_name', 'client_email', 'client_linkedin_url',
        'company_name', 'email_content_sent', 'opportunity_url', 'contact_number',
        'reason', 'next_step', 'cadence', 'client_whatsapp',
    ],
    'positive_responses': [
        'person_id', 'source_file_id', 'original_worksheet',
        'response_date', 'response_date_raw',
        'client_type', 'connected_date', 'lead_established_date', 'client_name',
        'client_linkedin_id', 'linkedin_id_associated', 'client_email',
        'client_contact_number', 'first_follow_up', 'num_follow_ups_taken',
        'num_gap_days', 'response_quality', 'client_first_revert', 'chat_summary', 'source',
    ],
    'leads_generated': [
        'person_id', 'source_file_id', 'original_worksheet',
        'inquiry_date', 'inquiry_date_raw',
        'lead_gen_executive', 'client_name', 'client_location', 'company_name',
        'company_size', 'client_designation', 'client_linkedin_url', 'client_email',
        'client_contact_number', 'summary', 'next_step', 'zoho_contact_link',
        'zoho_deal_link', 'lead_source', 'account', 'assigned_consultant',
        'current_status', 'lost_reason', 'meeting_done', 'prospect_qualified',
        'response', 'status',
    ],
    'data_extraction': [
        'person_id', 'source_file_id', 'original_worksheet',
        'activity_date', 'activity_date_raw', 'extraction_time',
        'lead_generation_exec', 'prospect_name', 'prospect_company',
        'client_email', 'client_email_2', 'client_linkedin_url',
        'opportunity_url', 'source_of_data', 'region', 'designation',
        'industry', 'contact_number', 'connection_request', 'email_sent', 'likely_usecase',
    ],
    'biddetail_tenders': [
        'person_id', 'source_file_id', 'original_worksheet',
        'serial_no', 'data_fetch_date', 'contract_date', 'query_name',
        'link_of_tender', 'tender_no', 'amount', 'contact_person_name',
        'company', 'contact_details', 'bid_details_email', 'linkedin_email', 'linkedin',
    ],
}

# ──────────────────────────────────────────────
# MAIN INGESTION
# ──────────────────────────────────────────────
def main():
    conn = psycopg2.connect(**DB)
    conn.autocommit = True
    cur = conn.cursor()

    # 1. Create schema
    print("=" * 70)
    print("STEP 1: Creating schema...")
    for stmt in SCHEMA_SQL.split(';'):
        s = stmt.strip()
        if s:
            cur.execute(s + ';')
    print("  Schema created (11 tables + indexes)")

    # 2. Insert persons
    print("\nSTEP 2: Inserting persons...")
    person_ids = {}
    for p in PERSONS:
        cur.execute(
            "INSERT INTO persons (full_name, short_name) VALUES (%s, %s) ON CONFLICT (full_name) DO UPDATE SET short_name=EXCLUDED.short_name RETURNING id",
            (p['full_name'], p['short_name'])
        )
        pid = cur.fetchone()[0]
        person_ids[p['full_name']] = pid
        print(f"  {p['full_name']} → person_id={pid}")

    # 3. Connect to Drive and download files
    print("\nSTEP 3: Connecting to Google Drive...")
    service = get_drive_service()

    # Download all files (export Google Sheets as xlsx)
    print("\nSTEP 4: Downloading files...")
    file_paths = {}
    for f in FILES:
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', f['name'])
        if not safe_name.endswith('.xlsx'):
            safe_name += '.xlsx'
        dest = os.path.join(DOWNLOAD_DIR, safe_name)
        if os.path.exists(dest):
            print(f"  Already downloaded: {safe_name}")
        else:
            print(f"  Downloading: {f['name']}...")
            try:
                download_file(service, f['drive_id'], dest)
                print(f"    OK ({os.path.getsize(dest):,} bytes)")
            except Exception as e:
                print(f"    ERROR: {e}")
                continue
        file_paths[f['drive_id']] = dest

    # 4. Register source files
    print("\nSTEP 5: Registering source files...")
    source_file_ids = {}
    for f in FILES:
        if f['drive_id'] not in file_paths:
            continue
        person_id = person_ids[f['person']]
        period_end = f['period_end']
        cur.execute("""
            INSERT INTO source_files (person_id, file_name, drive_file_id, file_type, period_start, period_end)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (drive_file_id) DO UPDATE SET file_name=EXCLUDED.file_name RETURNING id
        """, (person_id, f['name'], f['drive_id'], f['file_type'],
              f['period_start'], period_end))
        sf_id = cur.fetchone()[0]
        source_file_ids[f['drive_id']] = sf_id
        print(f"  {f['name'][:60]}... → source_file_id={sf_id}")

    # 5. Process each file
    print("\n" + "=" * 70)
    print("STEP 6: Processing worksheets...")
    print("=" * 70)

    total_sheets = 0
    total_rows_inserted = 0
    all_logs = []

    for f in FILES:
        if f['drive_id'] not in file_paths:
            continue
        path = file_paths[f['drive_id']]
        person_id = person_ids[f['person']]
        sf_id = source_file_ids[f['drive_id']]

        print(f"\n{'─' * 60}")
        print(f"FILE: {f['name'][:70]}")
        print(f"Person: {f['person']} | Type: {f['file_type']}")
        print(f"{'─' * 60}")

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERROR opening workbook: {e}")
            continue

        # Update total worksheets
        cur.execute("UPDATE source_files SET total_worksheets=%s WHERE id=%s", (len(wb.sheetnames), sf_id))

        for ws_name in wb.sheetnames:
            total_sheets += 1
            ws = wb[ws_name]
            headers_raw, data_rows = read_worksheet_data(ws)

            # Determine target table
            target_table = route_worksheet(ws_name)
            col_map = get_column_map(target_table)

            # Normalize headers
            headers_norm = [norm(h) for h in headers_raw]
            non_empty_headers = [h for h in headers_norm if h]

            print(f"\n  Worksheet: '{ws_name}'")
            print(f"  → Target: {target_table} | Headers: {len(non_empty_headers)} | Data rows: {len(data_rows)}")

            # Skip empty worksheets
            if len(data_rows) == 0:
                print(f"    SKIPPED (empty)")
                all_logs.append({
                    'sf_id': sf_id, 'person_id': person_id, 'ws': ws_name,
                    'table': target_table, 'rows_sheet': 0, 'rows_inserted': 0,
                    'cols_sheet': len(non_empty_headers), 'cols_mapped': 0,
                    'status': 'SKIPPED_EMPTY',
                })
                continue

            # Single-column tracker → other_worksheet_data
            if is_single_column_tracker(ws, headers_raw):
                print(f"    Single-column tracker → other_worksheet_data")
                target_table = 'other_worksheet_data'

            # Handle other_worksheet_data (JSONB storage)
            if target_table == 'other_worksheet_data':
                inserted = 0
                for row_num, row in enumerate(data_rows, start=2):
                    row_dict = {}
                    for i, val in enumerate(row):
                        key = str(headers_raw[i]).strip() if i < len(headers_raw) and headers_raw[i] else f'col_{i}'
                        if val is not None and str(val).strip():
                            row_dict[key] = str(val).strip()
                    if row_dict:
                        cur.execute(
                            "INSERT INTO other_worksheet_data (person_id, source_file_id, original_worksheet, row_number, row_data) VALUES (%s,%s,%s,%s,%s)",
                            (person_id, sf_id, ws_name, row_num, json.dumps(row_dict))
                        )
                        inserted += 1
                total_rows_inserted += inserted
                print(f"    Inserted {inserted} rows as JSONB")
                all_logs.append({
                    'sf_id': sf_id, 'person_id': person_id, 'ws': ws_name,
                    'table': 'other_worksheet_data', 'rows_sheet': len(data_rows), 'rows_inserted': inserted,
                    'cols_sheet': len(non_empty_headers), 'cols_mapped': len(non_empty_headers),
                    'status': 'SUCCESS' if inserted == len(data_rows) else 'PARTIAL',
                })
                continue

            # Check for Ragini leads with data-as-headers
            use_positional = is_ragini_leads_no_header(ws_name, f['person'], headers_raw)
            if use_positional:
                print(f"    SPECIAL: Ragini leads — positional mapping (first row is data, not headers)")
                all_data = [list(headers_raw)] + data_rows
                data_rows = all_data

            # Map and insert rows
            cols = TABLE_COLS[target_table]
            date_col_info = DATE_COL.get(target_table)
            inserted = 0
            skipped_empty = 0
            mapped_cols_set = set()
            batch = []

            for row in data_rows:
                if use_positional:
                    mapped = map_row_positional(row, RAGINI_LEADS_POSITIONAL)
                else:
                    skip = TARGET_TRACKING_SKIP if target_table == 'target_tracking' else set()
                    mapped = map_row(headers_raw, row, col_map, skip_cols=skip)

                if not mapped or all(v is None for v in mapped.values()):
                    skipped_empty += 1
                    continue

                # Parse date field
                if date_col_info:
                    date_col, raw_col = date_col_info
                    raw_val = mapped.get(date_col)
                    if raw_val is not None:
                        parsed_dt, raw_str = parse_date(raw_val)
                        mapped[date_col] = str(parsed_dt) if parsed_dt else None
                        mapped[raw_col] = raw_str
                    else:
                        mapped[raw_col] = None

                # Add provenance
                mapped['person_id'] = person_id
                mapped['source_file_id'] = sf_id
                mapped['original_worksheet'] = ws_name

                # Build tuple in column order
                values = tuple(mapped.get(c) for c in cols)
                batch.append(values)
                mapped_cols_set.update(k for k, v in mapped.items() if v is not None)

                if len(batch) >= 500:
                    placeholders = ','.join(['%s'] * len(cols))
                    col_str = ','.join(cols)
                    execute_values(cur,
                        f"INSERT INTO {target_table} ({col_str}) VALUES %s",
                        batch, template=f"({placeholders})")
                    inserted += len(batch)
                    batch = []

            if batch:
                placeholders = ','.join(['%s'] * len(cols))
                col_str = ','.join(cols)
                execute_values(cur,
                    f"INSERT INTO {target_table} ({col_str}) VALUES %s",
                    batch, template=f"({placeholders})")
                inserted += len(batch)

            total_rows_inserted += inserted
            data_count = len(data_rows) - skipped_empty
            status = 'SUCCESS' if inserted == data_count else ('PARTIAL' if inserted > 0 else 'FAILED')
            print(f"    Inserted {inserted}/{data_count} rows | Mapped cols: {len(mapped_cols_set)}")
            if skipped_empty > 0:
                print(f"    ({skipped_empty} fully empty rows excluded from count)")

            all_logs.append({
                'sf_id': sf_id, 'person_id': person_id, 'ws': ws_name,
                'table': target_table, 'rows_sheet': data_count, 'rows_inserted': inserted,
                'cols_sheet': len(non_empty_headers), 'cols_mapped': len(mapped_cols_set),
                'status': status,
            })

        wb.close()

    # 6. Write ingestion log
    print("\n" + "=" * 70)
    print("STEP 7: Writing ingestion log...")
    for log in all_logs:
        cur.execute("""
            INSERT INTO ingestion_log (source_file_id, person_id, worksheet_name, target_table,
                rows_in_sheet, rows_inserted, columns_in_sheet, columns_mapped, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (log['sf_id'], log['person_id'], log['ws'], log['table'],
              log['rows_sheet'], log['rows_inserted'], log['cols_sheet'], log['cols_mapped'], log['status']))

    # 7. VALIDATION
    print("\n" + "=" * 70)
    print("STEP 8: VALIDATION REPORT")
    print("=" * 70)

    # Layer 1: Row counts
    print("\n┌─ Layer 1: Row Count Match ─────────────────────────────────────────┐")
    cur.execute("""
        SELECT il.worksheet_name, sf.file_name, p.full_name, il.target_table,
               il.rows_in_sheet, il.rows_inserted, il.status
        FROM ingestion_log il
        JOIN source_files sf ON il.source_file_id = sf.id
        JOIN persons p ON il.person_id = p.id
        ORDER BY p.full_name, sf.file_name, il.worksheet_name
    """)
    all_pass = True
    for row in cur.fetchall():
        ws, fname, person, table, sheet_rows, db_rows, status = row
        match = '✅' if status in ('SUCCESS', 'SKIPPED_EMPTY') else '⚠️'
        if status not in ('SUCCESS', 'SKIPPED_EMPTY'):
            all_pass = False
        fname_short = fname[:45] + '...' if len(fname) > 45 else fname
        print(f"│ {match} {person:20s} | {ws:30s} | {table:25s} | {sheet_rows:5d} → {db_rows:5d} | {status}")
    layer1 = '✅ PASS' if all_pass else '⚠️ PARTIAL'
    print(f"└─ Layer 1 Result: {layer1} {'─' * 45}┘")

    # Layer 2: Table row counts
    print("\n┌─ Layer 2: Database Table Totals ──────────────────────────────────┐")
    tables = ['target_tracking', 'linkedin_connections', 'linkedin_followups',
              'linkedin_inmails', 'emails', 'positive_responses', 'leads_generated',
              'data_extraction', 'biddetail_tenders', 'other_worksheet_data']
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        cnt = cur.fetchone()[0]
        print(f"│  {t:30s} │ {cnt:8,d} rows")
    print(f"└{'─' * 68}┘")

    # Layer 3: Summary per person
    print("\n┌─ Layer 3: Per-Person Summary ─────────────────────────────────────┐")
    for p_name, p_id in person_ids.items():
        cur.execute("SELECT COUNT(*) FROM target_tracking WHERE person_id=%s", (p_id,))
        tt = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM linkedin_connections WHERE person_id=%s", (p_id,))
        lc = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM linkedin_followups WHERE person_id=%s", (p_id,))
        lf = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM linkedin_inmails WHERE person_id=%s", (p_id,))
        li = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM positive_responses WHERE person_id=%s", (p_id,))
        pr = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM leads_generated WHERE person_id=%s", (p_id,))
        lg = cur.fetchone()[0]
        total = tt + lc + lf + li + pr + lg
        print(f"│ {p_name:25s} │ TT:{tt:5d} LC:{lc:5d} LF:{lf:5d} LI:{li:5d} PR:{pr:4d} LG:{lg:4d} │ Total: {total:,d}")
    print(f"└{'─' * 68}┘")

    # Overall
    print(f"\n{'═' * 70}")
    print(f"  TOTAL WORKSHEETS: {total_sheets}")
    print(f"  TOTAL ROWS INSERTED: {total_rows_inserted:,d}")
    print(f"  INGESTION COMPLETE")
    print(f"{'═' * 70}")

    conn.close()

if __name__ == '__main__':
    main()
